from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django.contrib.auth.models import User

from django.http import HttpResponse
import csv
from django.core.mail import send_mail
from django.contrib import messages
import random, string
from django.conf import settings
<<<<<<< HEAD

from .models import FileUpload, StudentProfile
=======
from .utils import send_email_and_log

from .models import FileUpload, StudentProfile, EmailLog
>>>>>>> 851f815b28aefb73556c1cedd85f9d3afbb11056

# openpyxl is optional; fall back to CSV export if not installed
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False


@admin.register(FileUpload)
class FileUploadAdmin(admin.ModelAdmin):
<<<<<<< HEAD
    list_display = ('title', 'level', 'semester', 'category', 'uploaded_at', 'download_count', 'uploaded_by')
=======
    list_display = ('title', 'level', 'semester', 'category', 'uploaded_at', 'download_count')
>>>>>>> 851f815b28aefb73556c1cedd85f9d3afbb11056
    list_filter = ('level', 'semester', 'category', 'archived')
    actions = ('archive_selected',)
    search_fields = ('title',)

<<<<<<< HEAD
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        # Show only staff users in the uploaded_by dropdown
        if db_field.name == 'uploaded_by':
            kwargs['queryset'] = User.objects.filter(is_staff=True)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

=======
>>>>>>> 851f815b28aefb73556c1cedd85f9d3afbb11056
    def save_model(self, request, obj, form, change):
        if not change and getattr(obj, 'uploaded_by', None) is None:
            obj.uploaded_by = request.user
        super().save_model(request, obj, form, change)

    def archive_selected(self, request, queryset):
        updated = queryset.update(archived=True)
        self.message_user(request, f'{updated} file(s) marked archived.')
    archive_selected.short_description = 'Mark selected files as archived'


class StudentProfileInline(admin.StackedInline):
    model = StudentProfile
    can_delete = False
    verbose_name_plural = 'student profile'
    fk_name = 'user'
    fields = ('middle_name', 'level')


class UserAdmin(DjangoUserAdmin):
    inlines = (StudentProfileInline,)
    list_display = ('username', 'email', 'first_name', 'middle_name_display', 'last_name', 'level_display', 'is_staff')
    actions = ('make_staff', 'remove_staff', 'deactivate_users', 'activate_users', 'reset_passwords', 'export_selected_users_excel')

    def middle_name_display(self, obj):
        try:
            return obj.studentprofile.middle_name
        except StudentProfile.DoesNotExist:
            return ''
    middle_name_display.short_description = 'Middle name'

    def level_display(self, obj):
        try:
            return obj.studentprofile.level
        except StudentProfile.DoesNotExist:
            return ''
    level_display.short_description = 'Level'

    def make_staff(self, request, queryset):
        updated = queryset.update(is_staff=True)
        self.message_user(request, f'{updated} user(s) granted staff status.')
    make_staff.short_description = 'Make selected users staff'

    def remove_staff(self, request, queryset):
        updated = queryset.update(is_staff=False)
        self.message_user(request, f'{updated} user(s) removed staff status.')
    remove_staff.short_description = 'Remove staff status from selected users'

    def deactivate_users(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} user(s) deactivated.')
    deactivate_users.short_description = 'Deactivate selected users'

    def activate_users(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} user(s) activated.')
    activate_users.short_description = 'Activate selected users'

    def reset_passwords(self, request, queryset):
        for user in queryset:
            pwd = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
            user.set_password(pwd)
            user.save()
            if user.email:
<<<<<<< HEAD
                try:
                    res = send_mail(
                        'Your password has been reset',
                        f'Hello {user.username},\n\nYour password has been reset by an admin. Temporary password: {pwd}\nPlease login and change your password.',
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                    )
                    # do not log password-reset emails to EmailLog
                except Exception as exc:
                    # swallow/logging handled by server; do not create EmailLog entries
                    pass
=======
                # Use helper to send email and create EmailLog with detailed errors on failure.
                subject = 'Your password has been reset'
                body = f'Hello {user.username},\n\nYour password has been reset by an admin. Temporary password: {pwd}\nPlease login and change your password.'
                try:
                    ok = send_email_and_log(subject, body, settings.DEFAULT_FROM_EMAIL, [user.email])
                    # helper already records EmailLog; we only inform admin via messages
                except Exception as exc:
                    # As a last-resort fallback, record a failed EmailLog entry.
                    try:
                        EmailLog.objects.create(subject=subject, body=body, from_email=settings.DEFAULT_FROM_EMAIL, recipients=user.email, sent=False, error=str(exc))
                    except Exception:
                        pass
>>>>>>> 851f815b28aefb73556c1cedd85f9d3afbb11056
        self.message_user(request, 'Selected users have been reset and (if they have an email) notified.')
    reset_passwords.short_description = 'Reset password for selected users and email them'

    def export_selected_users_excel(self, request, queryset):
        # If openpyxl is not available, fall back to CSV
        if not OPENPYXL_AVAILABLE:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=users_export.csv'
            writer = csv.writer(response)
            headers = ['id', 'username', 'email', 'first_name', 'middle_name', 'last_name', 'is_active', 'is_staff']
            writer.writerow(headers)
            for user in queryset:
                try:
                    middle = user.studentprofile.middle_name
                except StudentProfile.DoesNotExist:
                    middle = ''
                row = [
                    user.pk,
                    user.username,
                    user.email,
                    user.first_name,
                    middle,
                    user.last_name,
                    'yes' if user.is_active else 'no',
                    'yes' if user.is_staff else 'no',
                ]
                writer.writerow(row)
            return response

        wb = Workbook()
        ws = wb.active
        headers = ['id', 'username', 'email', 'first_name', 'middle_name', 'last_name', 'is_active', 'is_staff']
        ws.append(headers)
        for user in queryset:
            try:
                middle = user.studentprofile.middle_name
            except StudentProfile.DoesNotExist:
                middle = ''
            row = [
                user.pk,
                user.username,
                user.email,
                user.first_name,
                middle,
                user.last_name,
                'yes' if user.is_active else 'no',
                'yes' if user.is_staff else 'no',
            ]
            ws.append(row)

        # Adjust column widths
        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(i)].width = min(50, (max_length + 2))

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users_export.xlsx'
        wb.save(response)
        return response
    export_selected_users_excel.short_description = 'Export selected users to Excel (.xlsx)'


# Register the custom User admin
try:
    admin.site.unregister(User)
except Exception:
    pass
admin.site.register(User, UserAdmin)
<<<<<<< HEAD
=======


@admin.register(EmailLog)
class EmailLogAdmin(admin.ModelAdmin):
    list_display = ('subject', 'recipients', 'sent', 'created_at')
    readonly_fields = ('subject', 'body', 'from_email', 'recipients', 'sent', 'error', 'created_at')
    search_fields = ('subject', 'recipients')
    ordering = ('-created_at',)
>>>>>>> 851f815b28aefb73556c1cedd85f9d3afbb11056
